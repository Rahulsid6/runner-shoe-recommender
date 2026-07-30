#!/usr/bin/env python3
"""Convert the editable Excel catalog intake workbook into data/shoes.json."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.catalog_quality import quality_report

DEFAULT_WORKBOOK = ROOT / "outputs/catalog-intake-template/runwise-india-shoe-catalog-intake.xlsx"
DEFAULT_OUTPUT = ROOT / "data/shoes.json"
HEADER_ROW = 5


def number(value: object) -> int | float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value) if float(value).is_integer() else float(value)
    return None


def text_list(value: object) -> list[str]:
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def text(value: object) -> str | None:
    return str(value).strip() or None if value is not None else None


def build_catalog(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    if "Catalog Intake" not in workbook.sheetnames:
        raise ValueError("Workbook must contain a 'Catalog Intake' sheet.")
    sheet = workbook["Catalog Intake"]
    headers = {str(cell.value).strip(): cell.column for cell in sheet[HEADER_ROW] if cell.value}
    required_headers = {"id", "brand", "model", "gender", "category", "MSRP (INR)", "official source URL"}
    missing = required_headers - set(headers)
    if missing:
        raise ValueError(f"Workbook is missing required columns: {', '.join(sorted(missing))}")

    shoes: list[dict[str, Any]] = []
    for row_cells in sheet.iter_rows(min_row=HEADER_ROW + 1):
        value = lambda header: row_cells[headers[header] - 1].value
        if not value("id"):
            continue
        scores = {name: number(value(label)) for name, label in {
            "cushioning": "cushioning", "responsiveness": "responsiveness", "stability": "stability score", "durability": "durability", "value": "value", "grip": "grip", "protection": "protection"
        }.items()}
        shoe = {
            "id": text(value("id")), "brand": text(value("brand")), "model": text(value("model")), "gender": text(value("gender")), "category": text(value("category")),
            "msrp_inr": number(value("MSRP (INR)")), "weight_g": number(value("weight (g)")), "drop_mm": number(value("drop (mm)")),
            "stack_mm_heel": number(value("stack heel (mm)")), "stack_mm_forefoot": number(value("stack forefoot (mm)")),
            "stability": text(value("stability")), "cushion": text(value("cushion")), "width_options": text_list(value("width options")),
            "best_use": text_list(value("best use")), "distance_focus": text_list(value("distance focus")), "scores": scores,
            "notes": [item.strip() for item in str(value("notes") or "").split(";") if item.strip()],
        }
        source_url = text(value("official source URL"))
        if source_url:
            shoe["source_url"] = source_url
        shoes.append(shoe)
    workbook.close()
    return {"schema_version": "1.0", "market": "IN", "currency": "INR", "last_updated": date.today().isoformat(), "notes": "Curated catalog imported from Excel intake workbook.", "shoes": shoes}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", nargs="?", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--write", action="store_true", help="Write the validated result to data/shoes.json (default is a dry run).")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output path used with --write.")
    args = parser.parse_args()
    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")
    catalog = build_catalog(args.workbook)
    report = quality_report(catalog)
    summary = report["summary"]
    print(f"Excel intake: {summary['variants']} variants | errors: {summary['errors']} | warnings: {summary['warnings']} | publish-ready: {summary['ready_to_publish']}")
    for issue in report["issues"]:
        print(f"{issue['severity'].upper()}: {issue['shoe_id']} · {issue['field']} — {issue['message']}")
    if summary["errors"]:
        raise SystemExit(1)
    if args.write:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(json.dumps(catalog, indent=2) + "\n")
        print(f"Wrote validated catalog to {args.output}")
    else:
        print("Dry run only. Use --write after reviewing the output.")


if __name__ == "__main__":
    main()
